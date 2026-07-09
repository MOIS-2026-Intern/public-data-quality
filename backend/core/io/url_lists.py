from __future__ import annotations

import csv
import re
from pathlib import Path

SUPPORTED_URL_LIST_SUFFIXES = {".txt", ".csv", ".tsv", ".xlsx", ".xls"}
URL_PATTERN = re.compile(r"https?://[^\s\"'<>]+", re.IGNORECASE)
TRAILING_URL_PUNCTUATION = ".,;:)]}>"
URL_LIST_HEADER_VALUES = {
    "url",
    "urls",
    "link",
    "links",
    "dataurl",
    "dataseturl",
    "fileurl",
    "downloadurl",
    "링크",
    "url링크",
    "다운로드url",
    "파일url",
}


def supported_url_list_suffixes_label() -> str:
    return ", ".join(sorted(SUPPORTED_URL_LIST_SUFFIXES))


def load_url_list(file_path: str | Path, *, strict: bool = False) -> list[str]:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".txt":
        urls = _strict_urls_from_text_file(path) if strict else _urls_from_text_file(path)
    elif suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        urls = (
            _strict_urls_from_delimited_file(path, delimiter)
            if strict
            else _urls_from_delimited_file(path, delimiter)
        )
    elif suffix == ".xlsx":
        urls = _strict_urls_from_xlsx(path) if strict else _urls_from_xlsx(path)
    elif suffix == ".xls":
        urls = _strict_urls_from_xls(path) if strict else _urls_from_xls(path)
    else:
        raise ValueError(
            f"지원하지 않는 URL 목록 파일 형식입니다: {suffix or '<none>'}. "
            f"지원 형식: {supported_url_list_suffixes_label()}"
        )

    if not urls:
        raise ValueError("업로드한 URL 목록 파일에서 http/https URL을 찾지 못했습니다.")
    return urls


def _append_unique(items: list[str], values: list[str]) -> None:
    seen = set(items)
    for value in values:
        if value in seen:
            continue
        items.append(value)
        seen.add(value)


def _extract_urls(text: str) -> list[str]:
    extracted: list[str] = []
    for match in URL_PATTERN.findall(text or ""):
        candidate = match.rstrip(TRAILING_URL_PUNCTUATION).strip()
        if candidate:
            extracted.append(candidate)
    return extracted


def _is_url_list_header(text: str) -> bool:
    normalized = re.sub(r"[\s_-]+", "", (text or "").strip().lower())
    return normalized in URL_LIST_HEADER_VALUES


def _contains_only_urls(text: str, urls: list[str]) -> bool:
    remainder = text.strip()
    for url in urls:
        remainder = remainder.replace(url, "", 1)
    return not remainder.strip(" \t\r\n,;:|[](){}<>\"'")


def _strict_urls_from_values(values: list[str]) -> list[str]:
    urls: list[str] = []
    for value in values:
        text = str(value or "").strip()
        if not text:
            continue
        if _is_url_list_header(text):
            continue
        extracted = _extract_urls(text)
        if not extracted or not _contains_only_urls(text, extracted):
            return []
        _append_unique(urls, extracted)
    return urls


def _urls_from_text_file(path: Path) -> list[str]:
    urls: list[str] = []
    for line in path.read_text(encoding="utf-8-sig", errors="ignore").splitlines():
        _append_unique(urls, _extract_urls(line))
    return urls


def _strict_urls_from_text_file(path: Path) -> list[str]:
    return _strict_urls_from_values(path.read_text(encoding="utf-8-sig", errors="ignore").splitlines())


def _urls_from_delimited_file(path: Path, delimiter: str) -> list[str]:
    urls: list[str] = []
    with path.open("r", encoding="utf-8-sig", errors="ignore", newline="") as handle:
        for row in csv.reader(handle, delimiter=delimiter):
            for cell in row:
                _append_unique(urls, _extract_urls(str(cell)))
    return urls


def _strict_urls_from_delimited_file(path: Path, delimiter: str) -> list[str]:
    values: list[str] = []
    with path.open("r", encoding="utf-8-sig", errors="ignore", newline="") as handle:
        for row in csv.reader(handle, delimiter=delimiter):
            values.extend(str(cell) for cell in row)
    return _strict_urls_from_values(values)


def _urls_from_xlsx(path: Path) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ValueError("XLSX URL 목록 업로드를 처리하려면 openpyxl이 필요합니다.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    urls: list[str] = []
    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                for cell in row or ():
                    if cell in (None, ""):
                        continue
                    _append_unique(urls, _extract_urls(str(cell)))
    finally:
        workbook.close()
    return urls


def _strict_urls_from_xlsx(path: Path) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ValueError("XLSX URL 목록 업로드를 처리하려면 openpyxl이 필요합니다.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    values: list[str] = []
    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                values.extend(str(cell) for cell in row or () if cell not in (None, ""))
    finally:
        workbook.close()
    return _strict_urls_from_values(values)


def _urls_from_xls(path: Path) -> list[str]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover
        raise ValueError("XLS URL 목록 업로드를 처리하려면 xlrd가 필요합니다.") from exc

    workbook = xlrd.open_workbook(path)
    urls: list[str] = []
    for sheet_index in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_index)
        for row_index in range(sheet.nrows):
            for value in sheet.row_values(row_index):
                if value in (None, ""):
                    continue
                _append_unique(urls, _extract_urls(str(value)))
    return urls


def _strict_urls_from_xls(path: Path) -> list[str]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover
        raise ValueError("XLS URL 목록 업로드를 처리하려면 xlrd가 필요합니다.") from exc

    workbook = xlrd.open_workbook(path)
    values: list[str] = []
    for sheet_index in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_index)
        for row_index in range(sheet.nrows):
            values.extend(str(value) for value in sheet.row_values(row_index) if value not in (None, ""))
    return _strict_urls_from_values(values)
