from __future__ import annotations

import csv
import json
from collections.abc import Iterator
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from ..config.constants import (
    UPLOAD_DATASET_ID_PREFIX,
    UPLOAD_DATASET_TYPE,
    UPLOAD_PROVIDER_CODE,
    UPLOAD_PROVIDER_NAME,
    UPLOAD_SERVICE_TYPE,
    UPLOAD_UPDATE_CYCLE,
)
from ..schema.models import DatasetMeta


def _stringify_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _clean_headers(values: list[object]) -> list[str]:
    return [_stringify_cell(value) for value in values]


def _csv_dialect(handle, fallback_delimiter: str = ","):
    sample = handle.read(8192)
    handle.seek(0)
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel()
        dialect.delimiter = fallback_delimiter
        return dialect


def _iter_delimited_rows(path: Path, fallback_delimiter: str = ",") -> Iterator[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, dialect=_csv_dialect(handle, fallback_delimiter))
        for row in reader:
            yield {str(key): (value or "") for key, value in row.items() if key is not None}


def _read_delimited_headers(path: Path, fallback_delimiter: str = ",") -> list[str]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, dialect=_csv_dialect(handle, fallback_delimiter))
        return [header.strip() for header in next(reader, []) if str(header).strip()]


def _iter_xlsx_rows(path: Path) -> Iterator[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ValueError("XLSX 업로드를 처리하려면 openpyxl이 필요합니다.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    row_iter = worksheet.iter_rows(values_only=True)
    header_row = next(row_iter, None)
    headers = _clean_headers(list(header_row or []))
    if not any(headers):
        workbook.close()
        raise ValueError("Uploaded dataset has no header row.")

    try:
        for row in row_iter:
            values = list(row or [])
            yield {
                header: _stringify_cell(values[index]) if index < len(values) else ""
                for index, header in enumerate(headers)
                if header
            }
    finally:
        workbook.close()


def _iter_xls_rows(path: Path) -> Iterator[dict[str, str]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover
        raise ValueError("XLS 업로드를 처리하려면 xlrd가 필요합니다.") from exc

    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(0)
    if sheet.nrows == 0:
        raise ValueError("Uploaded dataset has no header row.")

    headers = _clean_headers(sheet.row_values(0))
    if not any(headers):
        raise ValueError("Uploaded dataset has no header row.")

    for row_index in range(1, sheet.nrows):
        values = sheet.row_values(row_index)
        yield {
            header: _stringify_cell(values[index]) if index < len(values) else ""
            for index, header in enumerate(headers)
            if header
        }


RECORD_CONTAINER_KEYS = (
    "records",
    "record",
    "rows",
    "row",
    "items",
    "item",
    "data",
    "result",
    "results",
    "list",
    "body",
    "response",
)


def _flatten_mapping(value: dict[str, Any], prefix: str = "") -> dict[str, str]:
    row: dict[str, str] = {}
    for key, nested_value in value.items():
        clean_key = str(key).strip()
        if not clean_key:
            continue
        path_key = f"{prefix}.{clean_key}" if prefix else clean_key
        if isinstance(nested_value, dict):
            row.update(_flatten_mapping(nested_value, path_key))
        elif isinstance(nested_value, list):
            row[path_key] = json.dumps(nested_value, ensure_ascii=False) if nested_value else ""
        else:
            row[path_key] = _stringify_cell(nested_value)
    return row


def _rows_from_list(values: list[Any]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for value in values:
        if isinstance(value, dict):
            rows.append(_flatten_mapping(value))
        elif isinstance(value, list):
            rows.extend(_rows_from_list(value))
        else:
            rows.append({"value": _stringify_cell(value)})
    return rows


def _find_records(value: Any) -> Any:
    if isinstance(value, list):
        return value
    if not isinstance(value, dict):
        return value

    for key in RECORD_CONTAINER_KEYS:
        if key in value:
            candidate = _find_records(value[key])
            if isinstance(candidate, list) and candidate:
                return candidate

    queue = list(value.values())
    while queue:
        candidate = queue.pop(0)
        if isinstance(candidate, list) and candidate and any(isinstance(item, dict) for item in candidate):
            return candidate
        if isinstance(candidate, dict):
            queue.extend(candidate.values())
    return value


def _json_payload_rows(payload: Any) -> list[dict[str, str]]:
    records = _find_records(payload)
    if isinstance(records, list):
        return _rows_from_list(records)
    if isinstance(records, dict):
        return [_flatten_mapping(records)]
    return [{"value": _stringify_cell(records)}]


def _iter_json_rows(path: Path) -> Iterator[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig") as handle:
        payload = json.load(handle)
    yield from _json_payload_rows(payload)


def _iter_jsonl_rows(path: Path) -> Iterator[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig") as handle:
        for line in handle:
            if not line.strip():
                continue
            payload = json.loads(line)
            yield from _json_payload_rows(payload)


def _xml_tag_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def _flatten_xml_element(element: ElementTree.Element, prefix: str = "") -> dict[str, str]:
    row: dict[str, str] = {}
    for key, value in element.attrib.items():
        attr_key = f"{prefix}.@{_xml_tag_name(key)}" if prefix else f"@{_xml_tag_name(key)}"
        row[attr_key] = _stringify_cell(value)

    children = list(element)
    if not children:
        text = _stringify_cell(element.text)
        if prefix and text:
            row[prefix] = text
        return row

    tag_counts: dict[str, int] = {}
    for child in children:
        child_tag = _xml_tag_name(child.tag)
        tag_counts[child_tag] = tag_counts.get(child_tag, 0) + 1

    for child in children:
        child_tag = _xml_tag_name(child.tag)
        child_key = f"{prefix}.{child_tag}" if prefix else child_tag
        if tag_counts[child_tag] > 1 and not list(child):
            current = row.get(child_key)
            child_text = _stringify_cell(child.text)
            row[child_key] = child_text if current is None else f"{current}, {child_text}"
        else:
            row.update(_flatten_xml_element(child, child_key))
    return row


def _xml_record_elements(root: ElementTree.Element) -> list[ElementTree.Element]:
    preferred_tags = {"item", "row", "record", "data"}
    best: list[ElementTree.Element] = []
    best_score = -1

    for parent in root.iter():
        groups: dict[str, list[ElementTree.Element]] = {}
        for child in list(parent):
            groups.setdefault(_xml_tag_name(child.tag).lower(), []).append(child)

        for tag, elements in groups.items():
            if len(elements) < 2:
                continue
            score = len(elements) * 10 + (1000 if tag in preferred_tags else 0)
            if score > best_score:
                best = elements
                best_score = score

    if best:
        return best

    for element in root.iter():
        if _xml_tag_name(element.tag).lower() in preferred_tags and list(element):
            best.append(element)
    return best or [root]


def _iter_xml_rows(path: Path) -> Iterator[dict[str, str]]:
    root = ElementTree.parse(path).getroot()
    for element in _xml_record_elements(root):
        row = _flatten_xml_element(element)
        if row:
            yield row


def _headers_from_rows(rows: list[dict[str, str]]) -> list[str]:
    headers: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for header in row:
            if header not in seen:
                headers.append(header)
                seen.add(header)
    return headers


def iter_uploaded_rows(file_path: str | Path) -> Iterator[dict[str, str]]:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        yield from _iter_delimited_rows(path, ",")
        return
    if suffix == ".tsv":
        yield from _iter_delimited_rows(path, "\t")
        return
    if suffix == ".txt":
        yield from _iter_delimited_rows(path, ",")
        return
    if suffix == ".xlsx":
        yield from _iter_xlsx_rows(path)
        return
    if suffix == ".xls":
        yield from _iter_xls_rows(path)
        return
    if suffix == ".json":
        yield from _iter_json_rows(path)
        return
    if suffix == ".jsonl":
        yield from _iter_jsonl_rows(path)
        return
    if suffix == ".xml":
        yield from _iter_xml_rows(path)
        return
    raise ValueError(f"Unsupported file type: {suffix or '<none>'}. Supported: .csv, .tsv, .txt, .xlsx, .xls, .json, .jsonl, .xml")


def load_uploaded_headers(file_path: str | Path) -> list[str]:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_delimited_headers(path, ",")
    if suffix == ".tsv":
        return _read_delimited_headers(path, "\t")
    if suffix == ".txt":
        return _read_delimited_headers(path, ",")
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise ValueError("XLSX 업로드를 처리하려면 openpyxl이 필요합니다.") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        header_row = next(worksheet.iter_rows(values_only=True), None)
        workbook.close()
        return [header for header in _clean_headers(list(header_row or [])) if header]
    if suffix == ".xls":
        try:
            import xlrd
        except ImportError as exc:  # pragma: no cover
            raise ValueError("XLS 업로드를 처리하려면 xlrd가 필요합니다.") from exc
        workbook = xlrd.open_workbook(path)
        sheet = workbook.sheet_by_index(0)
        if sheet.nrows == 0:
            return []
        return [header for header in _clean_headers(sheet.row_values(0)) if header]
    if suffix == ".json":
        return _headers_from_rows(list(_iter_json_rows(path)))
    if suffix == ".jsonl":
        return _headers_from_rows(list(_iter_jsonl_rows(path)))
    if suffix == ".xml":
        return _headers_from_rows(list(_iter_xml_rows(path)))
    raise ValueError(f"Unsupported file type: {suffix or '<none>'}. Supported: .csv, .tsv, .txt, .xlsx, .xls, .json, .jsonl, .xml")


def _split_csv_list(value: str) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in value.split(",") if item.strip()]


def _file_cache_key(file_path: str | Path) -> tuple[str, int]:
    path = Path(file_path).resolve()
    return str(path), path.stat().st_mtime_ns


def load_dataset_meta(
    csv_path: str | Path,
    dataset_id: str | None = None,
    dataset_name: str | None = None,
) -> DatasetMeta:
    path = Path(csv_path)
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            row_id = (row.get("목록키") or "").strip()
            row_name = (row.get("목록명") or "").strip()
            if dataset_id and row_id != dataset_id:
                continue
            if dataset_name and row_name != dataset_name:
                continue
            total_rows = (row.get("전체행") or "").strip()
            return DatasetMeta(
                dataset_id=row_id,
                dataset_name=row_name,
                keywords=_split_csv_list(row.get("키워드", "")),
                provider_name=(row.get("제공기관명") or "").strip(),
                provider_code=(row.get("제공기관코드") or "").strip(),
                dataset_type=(row.get("목록유형") or "").strip(),
                service_type=(row.get("서비스 유형") or "").strip(),
                data_format=(row.get("데이터포맷") or "").strip(),
                request_fields=_split_csv_list(row.get("요청변수", "")),
                response_fields=_split_csv_list(row.get("출력결과", "")),
                update_cycle=(row.get("주기") or "").strip(),
                total_rows=int(total_rows) if total_rows.isdigit() else None,
            )
    target = dataset_id or dataset_name or "<unknown>"
    raise ValueError(f"Dataset not found: {target}")


def load_uploaded_dataset_meta(file_path: str | Path, dataset_name: str | None = None) -> DatasetMeta:
    path = Path(file_path)
    suffix = path.suffix.lstrip(".").lower() or "csv"
    header = load_uploaded_headers(path)
    if not header:
        raise ValueError("Uploaded dataset has no header row.")

    name = dataset_name or path.stem
    return DatasetMeta(
        dataset_id=f"{UPLOAD_DATASET_ID_PREFIX}{path.stem}",
        dataset_name=name,
        keywords=[],
        provider_name=UPLOAD_PROVIDER_NAME,
        provider_code=UPLOAD_PROVIDER_CODE,
        dataset_type=UPLOAD_DATASET_TYPE,
        service_type=UPLOAD_SERVICE_TYPE,
        data_format=suffix,
        request_fields=[],
        response_fields=[column.strip() for column in header if column.strip()],
        update_cycle=UPLOAD_UPDATE_CYCLE,
        total_rows=None,
    )
