from __future__ import annotations

import re
import uuid
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPORTS_DIR_NAME = "reports"
ERROR_REPORT_SUFFIX = "error_report.xlsx"
MAX_COMMENT_LENGTH = 32000


HEADER_FILL = PatternFill("solid", fgColor="EAF0F6")
ISSUE_FILL = PatternFill("solid", fgColor="FBE3DF")
WARNING_FILL = PatternFill("solid", fgColor="FFF1CC")
SUMMARY_FILL = PatternFill("solid", fgColor="F4F7FA")
HEADER_FONT = Font(bold=True, color="10221B")


def write_error_report(
    *,
    result: dict[str, Any],
    validation_rows: list[dict[str, str]],
    output_dir: Path,
) -> Path:
    reports_dir = output_dir / REPORTS_DIR_NAME
    reports_dir.mkdir(parents=True, exist_ok=True)

    dataset_name = str(result.get("summary", {}).get("dataset_name") or "dataset")
    report_path = reports_dir / _report_filename(dataset_name)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "요약"
    _write_summary_sheet(summary_sheet, result, validation_rows)
    _write_data_sheet(workbook.create_sheet("전체 데이터 오류 표시"), result, validation_rows)
    _write_findings_sheet(workbook.create_sheet("오류 상세"), result, validation_rows)
    _write_column_stats_sheet(workbook.create_sheet("컬럼별 오류 통계"), result)

    workbook.save(report_path)
    return report_path


def _report_filename(dataset_name: str) -> str:
    safe_name = re.sub(r"[^0-9A-Za-z가-힣._-]+", "_", dataset_name).strip("._") or "dataset"
    safe_name = safe_name[:80]
    return f"{safe_name}_{uuid.uuid4().hex[:8]}_{ERROR_REPORT_SUFFIX}"


def _write_summary_sheet(sheet, result: dict[str, Any], validation_rows: list[dict[str, str]]) -> None:
    summary = result.get("summary", {})
    issue_cells = _issue_cells_by_position(result)
    issue_rows = {row_index for row_index, _ in issue_cells}
    issue_columns = {column_name for _, column_name in issue_cells}
    rows = [
        ("데이터셋", summary.get("dataset_name", "")),
        ("제공기관", summary.get("provider_name", "")),
        ("전체 행 수", summary.get("row_count") or len(validation_rows)),
        ("컬럼 수", summary.get("column_count", 0)),
        ("오류 건수", summary.get("issue_finding_count", 0)),
        ("오류 발생 행 수", len(issue_rows)),
        ("오류 발생 컬럼 수", len(issue_columns)),
        ("수정 제안 수", summary.get("repair_suggestion_count", 0)),
    ]
    sheet.append(["항목", "값"])
    for row in rows:
        sheet.append(list(row))
    _style_header_row(sheet, 1)
    _auto_width(sheet)
    sheet.freeze_panes = "A2"


def _write_data_sheet(sheet, result: dict[str, Any], validation_rows: list[dict[str, str]]) -> None:
    headers = _result_headers(result, validation_rows)
    sheet.append(["row_index", *headers])
    _style_header_row(sheet, 1)
    sheet.freeze_panes = "A2"

    issue_map = _issue_messages_by_cell(result)
    max_finding_row = max((row_index for row_index, _ in issue_map), default=0)
    row_count = max(len(validation_rows), max_finding_row)

    for row_index in range(1, row_count + 1):
        source_row = validation_rows[row_index - 1] if row_index <= len(validation_rows) else {}
        sheet.append([row_index, *[source_row.get(header, "") for header in headers]])

    column_index_by_name = {header: index + 2 for index, header in enumerate(headers)}
    for (row_index, column_name), messages in issue_map.items():
        excel_row = row_index + 1
        excel_column = column_index_by_name.get(column_name)
        if not excel_column:
            continue
        cell = sheet.cell(row=excel_row, column=excel_column)
        cell.fill = ISSUE_FILL
        cell.comment = Comment(_comment_text(messages), "LDQ")

    _auto_width(sheet, max_width=42)


def _write_findings_sheet(sheet, result: dict[str, Any], validation_rows: list[dict[str, str]]) -> None:
    headers = [
        "행 번호",
        "컬럼명",
        "현재 값",
        "심각도",
        "검증 영역",
        "검증 기준",
        "규칙",
        "오류 메시지",
        "LLM 최종 검증",
        "관련 컬럼",
        "근거",
    ]
    sheet.append(headers)
    _style_header_row(sheet, 1)
    sheet.freeze_panes = "A2"

    for finding in result.get("findings", []):
        if finding.get("finding_type") != "issue":
            continue
        column_name = str(finding.get("column_name") or "")
        row_indexes = _finding_row_indexes(finding)
        if not row_indexes:
            row_indexes = [None]
        for row_index in row_indexes:
            current_value = ""
            if row_index and 0 < row_index <= len(validation_rows):
                current_value = validation_rows[row_index - 1].get(column_name, "")
            sheet.append(
                [
                    row_index or "",
                    column_name,
                    current_value,
                    finding.get("severity", ""),
                    finding.get("category_label", ""),
                    finding.get("criterion_name", ""),
                    finding.get("rule_id", ""),
                    finding.get("message", ""),
                    finding.get("llm_final_verification", ""),
                    ", ".join(finding.get("related_columns") or []),
                    " | ".join(finding.get("evidence") or []),
                ]
            )

    _auto_width(sheet, max_width=48)


def _write_column_stats_sheet(sheet, result: dict[str, Any]) -> None:
    counter: Counter[str] = Counter()
    severe_counter: dict[str, Counter[str]] = defaultdict(Counter)
    for finding in result.get("findings", []):
        if finding.get("finding_type") != "issue":
            continue
        column_name = str(finding.get("column_name") or "")
        count = len(_finding_row_indexes(finding)) or 1
        counter[column_name] += count
        severe_counter[column_name][str(finding.get("severity") or "")] += count

    sheet.append(["컬럼명", "오류 건수", "error", "warning", "info"])
    _style_header_row(sheet, 1)
    for column_name, count in counter.most_common():
        sheet.append(
            [
                column_name,
                count,
                severe_counter[column_name].get("error", 0),
                severe_counter[column_name].get("warning", 0),
                severe_counter[column_name].get("info", 0),
            ]
        )
    _auto_width(sheet)


def _result_headers(result: dict[str, Any], validation_rows: list[dict[str, str]]) -> list[str]:
    headers = [str(header) for header in result.get("preview_headers") or [] if str(header)]
    seen = set(headers)
    for row in validation_rows:
        for header in row:
            if header not in seen:
                headers.append(header)
                seen.add(header)
    return headers


def _finding_row_indexes(finding: dict[str, Any]) -> list[int]:
    indexes: list[int] = []
    for row_index in finding.get("row_indexes") or []:
        try:
            parsed = int(row_index)
        except (TypeError, ValueError):
            continue
        if parsed > 0:
            indexes.append(parsed)
    return indexes


def _issue_cells_by_position(result: dict[str, Any]) -> set[tuple[int, str]]:
    return set(_issue_messages_by_cell(result))


def _issue_messages_by_cell(result: dict[str, Any]) -> dict[tuple[int, str], list[str]]:
    issue_map: dict[tuple[int, str], list[str]] = defaultdict(list)
    for finding in result.get("findings", []):
        if finding.get("finding_type") != "issue":
            continue
        column_name = str(finding.get("column_name") or "")
        if not column_name:
            continue
        message = _finding_message(finding)
        for row_index in _finding_row_indexes(finding):
            issue_map[(row_index, column_name)].append(message)
    return issue_map


def _finding_message(finding: dict[str, Any]) -> str:
    parts = [
        str(finding.get("category_label") or "").strip(),
        str(finding.get("criterion_name") or "").strip(),
        str(finding.get("message") or "").strip(),
    ]
    llm_final_verification = str(finding.get("llm_final_verification") or "").strip()
    if llm_final_verification:
        parts.append(f"LLM 최종 검증: {llm_final_verification}")
    rule_id = str(finding.get("rule_id") or "").strip()
    message = " / ".join(part for part in parts if part)
    return f"{message} ({rule_id})" if rule_id else message


def _comment_text(messages: list[str]) -> str:
    text = "\n\n".join(f"- {message}" for message in messages if message)
    return text[:MAX_COMMENT_LENGTH]


def _style_header_row(sheet, row_index: int) -> None:
    for cell in sheet[row_index]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")


def _auto_width(sheet, max_width: int = 36) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        width = max(
            min(len(str(cell.value or "")) + 2, max_width)
            for cell in column_cells
        )
        sheet.column_dimensions[column_letter].width = max(width, 10)
