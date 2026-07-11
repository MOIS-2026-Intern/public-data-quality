from __future__ import annotations

import csv
import json
import re
from collections import Counter
from collections.abc import Iterator
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from backend.config.uploads import (
    UPLOAD_DATASET_ID_PREFIX,
    UPLOAD_DATASET_TYPE,
    UPLOAD_PROVIDER_CODE,
    UPLOAD_PROVIDER_NAME,
    UPLOAD_SERVICE_TYPE,
    UPLOAD_UPDATE_CYCLE,
)
from backend.domain.entities.models import DatasetMeta
from .text_encoding import detect_text_encoding


def _stringify_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _clean_headers(values: list[object]) -> list[str]:
    return [_stringify_cell(value) for value in values]


def _detect_text_encoding(path: Path) -> str:
    return detect_text_encoding(
        path.read_bytes(),
        error_message="텍스트 데이터 파일 인코딩을 판별할 수 없습니다. UTF-8 또는 CP949/EUC-KR CSV로 저장해 다시 시도하세요.",
    )


def _open_text_dataset(path: Path):
    return path.open("r", encoding=_detect_text_encoding(path), newline="")


def _csv_dialect(handle, fallback_delimiter: str = ","):
    sample = handle.read(8192)
    handle.seek(0)
    delimiter = _consistent_delimiter(sample, fallback_delimiter)
    if delimiter:
        dialect = csv.excel()
        dialect.delimiter = delimiter
        return dialect
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel()
        dialect.delimiter = fallback_delimiter
        return dialect


def _consistent_delimiter(sample: str, fallback_delimiter: str = ",") -> str | None:
    lines = [line for line in sample.splitlines() if line.strip()][:50]
    if len(lines) < 2:
        return None

    candidates = ["\t", fallback_delimiter, ",", ";", "|"]
    best: tuple[float, str] | None = None
    for delimiter in dict.fromkeys(candidates):
        counts = [line.count(delimiter) for line in lines]
        positive_counts = [count for count in counts if count > 0]
        if len(positive_counts) < max(2, int(len(lines) * 0.7)):
            continue

        common_count, common_frequency = Counter(positive_counts).most_common(1)[0]
        consistency = common_frequency / len(positive_counts)
        if consistency < 0.8:
            continue

        coverage = len(positive_counts) / len(lines)
        delimiter_bias = 0.5 if delimiter == "\t" else 0.1 if delimiter == fallback_delimiter else 0
        score = coverage * 10 + consistency * 10 + min(common_count, 20) * 0.1 + delimiter_bias
        if best is None or score > best[0]:
            best = (score, delimiter)
    return best[1] if best else None


def _row_mapping(headers: list[str], values: list[object], delimiter: str) -> dict[str, str]:
    cleaned_headers = [header for header in _clean_headers(headers) if header]
    cleaned_values = [_stringify_cell(value) for value in values]
    repaired_values = _repair_overflow_row(cleaned_headers, cleaned_values, delimiter)
    return {
        header: repaired_values[index] if index < len(repaired_values) else ""
        for index, header in enumerate(cleaned_headers)
    }


def _repair_overflow_row(headers: list[str], values: list[str], delimiter: str) -> list[str]:
    if not headers:
        return []
    if len(values) <= len(headers):
        return values + [""] * (len(headers) - len(values))

    repaired = [""] * len(headers)
    left_header_index = 0
    left_value_index = 0
    while (
        left_header_index < len(headers)
        and left_value_index < len(values)
        and _is_anchor_header(headers[left_header_index])
        and _value_matches_header(headers[left_header_index], values[left_value_index])
    ):
        repaired[left_header_index] = values[left_value_index]
        left_header_index += 1
        left_value_index += 1

    right_header_index = len(headers) - 1
    right_value_index = len(values) - 1
    right_anchor_count = 0
    while (
        right_header_index >= left_header_index
        and right_value_index >= left_value_index
        and _is_anchor_header(headers[right_header_index])
        and _value_matches_header(headers[right_header_index], values[right_value_index])
    ):
        repaired[right_header_index] = values[right_value_index]
        right_header_index -= 1
        right_value_index -= 1
        right_anchor_count += 1

    if right_anchor_count == 0:
        return _merge_overflow_values(headers, values, delimiter)

    middle_headers = headers[left_header_index : right_header_index + 1]
    middle_values = values[left_value_index : right_value_index + 1]
    middle_repaired = _merge_overflow_values(middle_headers, middle_values, delimiter)
    for offset, value in enumerate(middle_repaired):
        repaired[left_header_index + offset] = value
    return repaired


def _merge_overflow_values(headers: list[str], values: list[str], delimiter: str) -> list[str]:
    if not headers:
        return []
    if len(values) <= len(headers):
        return values + [""] * (len(headers) - len(values))

    merge_index = _overflow_merge_index(headers)
    repaired: list[str] = []
    value_index = 0
    for header_index in range(len(headers)):
        remaining_headers = len(headers) - header_index - 1
        if header_index == merge_index:
            take_count = max(1, len(values) - value_index - remaining_headers)
            repaired.append(_join_overflow_values(values[value_index : value_index + take_count], delimiter))
            value_index += take_count
            continue
        repaired.append(values[value_index] if value_index < len(values) else "")
        value_index += 1
    return repaired


def _join_overflow_values(values: list[str], delimiter: str) -> str:
    joiner = ", " if delimiter == "," else delimiter
    return joiner.join(values)


def _overflow_merge_index(headers: list[str]) -> int:
    for index, header in enumerate(headers):
        if _is_free_text_header(header):
            return index
    return 0


def _normalized_header_name(header: str) -> str:
    return re.sub(r"[\s_-]+", "", str(header or "").strip().lower())


def _is_free_text_header(header: str) -> bool:
    name = _normalized_header_name(header)
    return any(
        marker in name
        for marker in (
            "내용",
            "사유",
            "이유",
            "개요",
            "설명",
            "비고",
            "메모",
            "상세",
            "본문",
            "텍스트",
            "제안",
            "name",
            "description",
            "desc",
            "content",
            "summary",
            "reason",
            "memo",
            "note",
        )
    )


def _anchor_kind(header: str) -> str | None:
    name = _normalized_header_name(header)
    if any(marker in name for marker in ("url", "링크", "홈페이지", "사이트")):
        return "url"
    if name in {"연도", "년도", "year"} or name.endswith("연도") or name.endswith("년도"):
        return "year"
    if any(marker in name for marker in ("일자", "날짜", "일시", "기준일", "시작일", "종료일", "date")) or name.endswith("일"):
        return "date"
    if any(marker in name for marker in ("연번", "순번", "코드", "번호", "식별자", "id", "key")):
        return "code"
    if (
        name.endswith("수")
        or any(marker in name for marker in ("건수", "횟수", "인원", "금액", "가격", "비율", "면적", "거리", "count", "amount", "number"))
    ):
        return "number"
    return None


def _is_anchor_header(header: str) -> bool:
    return _anchor_kind(header) is not None


def _value_matches_header(header: str, value: str) -> bool:
    text = str(value or "").strip()
    if not text:
        return False

    kind = _anchor_kind(header)
    if kind == "url":
        return bool(re.fullmatch(r"https?://\S+", text, flags=re.IGNORECASE))
    if kind == "year":
        return bool(re.fullmatch(r"\d{4}", text))
    if kind == "date":
        return bool(
            re.fullmatch(r"\d{4}[-/.]\d{1,2}[-/.]\d{1,2}(?:\s*\([^)]+\))?", text)
            or re.fullmatch(r"\d{2}[.]\d{1,2}[.]\d{1,2}[.]?", text)
        )
    if kind == "number":
        return bool(re.fullmatch(r"-?\d{1,3}(?:,\d{3})*(?:\.\d+)?%?|-?\d+(?:\.\d+)?%?", text))
    if kind == "code":
        return bool(re.fullmatch(r"[0-9A-Za-z가-힣._:-]{1,80}", text))
    return False


def _iter_delimited_rows(path: Path, fallback_delimiter: str = ",") -> Iterator[dict[str, str]]:
    with _open_text_dataset(path) as handle:
        dialect = _csv_dialect(handle, fallback_delimiter)
        reader = csv.reader(handle, dialect=dialect)
        headers = _clean_headers(next(reader, []))
        for row in reader:
            yield _row_mapping(headers, row, dialect.delimiter)


def _read_delimited_headers(path: Path, fallback_delimiter: str = ",") -> list[str]:
    with _open_text_dataset(path) as handle:
        reader = csv.reader(handle, dialect=_csv_dialect(handle, fallback_delimiter))
        return [header.strip() for header in next(reader, []) if str(header).strip()]


def _iter_xlsx_rows(path: Path) -> Iterator[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ValueError("XLSX 업로드를 처리하려면 openpyxl이 필요합니다.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    row_iter = worksheet.iter_rows(values_only=True)
    header_row = next(row_iter, None)
    headers = _clean_headers(list(header_row or []))
    if not any(headers):
        workbook.close()
        raise ValueError("Uploaded dataset has no header row.")

    try:
        for row in row_iter:
            values = list(row or [])
            yield {
                header: _stringify_cell(values[index]) if index < len(values) else ""
                for index, header in enumerate(headers)
                if header
            }
    finally:
        workbook.close()


def _iter_xls_rows(path: Path) -> Iterator[dict[str, str]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover
        raise ValueError("XLS 업로드를 처리하려면 xlrd가 필요합니다.") from exc

    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(0)
    if sheet.nrows == 0:
        raise ValueError("Uploaded dataset has no header row.")

    headers = _clean_headers(sheet.row_values(0))
    if not any(headers):
        raise ValueError("Uploaded dataset has no header row.")

    for row_index in range(1, sheet.nrows):
        values = sheet.row_values(row_index)
        yield {
            header: _stringify_cell(values[index]) if index < len(values) else ""
            for index, header in enumerate(headers)
            if header
        }


RECORD_CONTAINER_KEYS = (
    "records",
    "record",
    "rows",
    "row",
    "items",
    "item",
    "data",
    "result",
    "results",
    "list",
    "body",
    "response",
)


def _flatten_mapping(value: dict[str, Any], prefix: str = "") -> dict[str, str]:
    row: dict[str, str] = {}
    for key, nested_value in value.items():
        clean_key = str(key).strip()
        if not clean_key:
            continue
        path_key = f"{prefix}.{clean_key}" if prefix else clean_key
        if isinstance(nested_value, dict):
            row.update(_flatten_mapping(nested_value, path_key))
        elif isinstance(nested_value, list):
            row[path_key] = json.dumps(nested_value, ensure_ascii=False) if nested_value else ""
        else:
            row[path_key] = _stringify_cell(nested_value)
    return row


def _rows_from_list(values: list[Any]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for value in values:
        if isinstance(value, dict):
            rows.append(_flatten_mapping(value))
        elif isinstance(value, list):
            rows.extend(_rows_from_list(value))
        else:
            rows.append({"value": _stringify_cell(value)})
    return rows


def _find_records(value: Any) -> Any:
    if isinstance(value, list):
        return value
    if not isinstance(value, dict):
        return value

    for key in RECORD_CONTAINER_KEYS:
        if key in value:
            candidate = _find_records(value[key])
            if isinstance(candidate, list) and candidate:
                return candidate

    queue = list(value.values())
    while queue:
        candidate = queue.pop(0)
        if isinstance(candidate, list) and candidate and any(isinstance(item, dict) for item in candidate):
            return candidate
        if isinstance(candidate, dict):
            queue.extend(candidate.values())
    return value


def _json_payload_rows(payload: Any) -> list[dict[str, str]]:
    records = _find_records(payload)
    if isinstance(records, list):
        return _rows_from_list(records)
    if isinstance(records, dict):
        return [_flatten_mapping(records)]
    return [{"value": _stringify_cell(records)}]


def _iter_json_rows(path: Path) -> Iterator[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig") as handle:
        payload = json.load(handle)
    yield from _json_payload_rows(payload)


def _iter_jsonl_rows(path: Path) -> Iterator[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig") as handle:
        for line in handle:
            if not line.strip():
                continue
            payload = json.loads(line)
            yield from _json_payload_rows(payload)


def _xml_tag_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def _flatten_xml_element(element: ElementTree.Element, prefix: str = "") -> dict[str, str]:
    row: dict[str, str] = {}
    for key, value in element.attrib.items():
        attr_key = f"{prefix}.@{_xml_tag_name(key)}" if prefix else f"@{_xml_tag_name(key)}"
        row[attr_key] = _stringify_cell(value)

    children = list(element)
    if not children:
        text = _stringify_cell(element.text)
        if prefix and text:
            row[prefix] = text
        return row

    tag_counts: dict[str, int] = {}
    for child in children:
        child_tag = _xml_tag_name(child.tag)
        tag_counts[child_tag] = tag_counts.get(child_tag, 0) + 1

    for child in children:
        child_tag = _xml_tag_name(child.tag)
        child_key = f"{prefix}.{child_tag}" if prefix else child_tag
        if tag_counts[child_tag] > 1 and not list(child):
            current = row.get(child_key)
            child_text = _stringify_cell(child.text)
            row[child_key] = child_text if current is None else f"{current}, {child_text}"
        else:
            row.update(_flatten_xml_element(child, child_key))
    return row


def _xml_record_elements(root: ElementTree.Element) -> list[ElementTree.Element]:
    preferred_tags = {"item", "row", "record", "data"}
    best: list[ElementTree.Element] = []
    best_score = -1

    for parent in root.iter():
        groups: dict[str, list[ElementTree.Element]] = {}
        for child in list(parent):
            groups.setdefault(_xml_tag_name(child.tag).lower(), []).append(child)

        for tag, elements in groups.items():
            if len(elements) < 2:
                continue
            score = len(elements) * 10 + (1000 if tag in preferred_tags else 0)
            if score > best_score:
                best = elements
                best_score = score

    if best:
        return best

    for element in root.iter():
        if _xml_tag_name(element.tag).lower() in preferred_tags and list(element):
            best.append(element)
    return best or [root]


def _iter_xml_rows(path: Path) -> Iterator[dict[str, str]]:
    root = ElementTree.parse(path).getroot()
    for element in _xml_record_elements(root):
        row = _flatten_xml_element(element)
        if row:
            yield row


def _headers_from_rows(rows: list[dict[str, str]]) -> list[str]:
    headers: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for header in row:
            if header not in seen:
                headers.append(header)
                seen.add(header)
    return headers


def iter_uploaded_rows(file_path: str | Path) -> Iterator[dict[str, str]]:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        yield from _iter_delimited_rows(path, ",")
        return
    if suffix == ".tsv":
        yield from _iter_delimited_rows(path, "\t")
        return
    if suffix == ".txt":
        yield from _iter_delimited_rows(path, ",")
        return
    if suffix == ".xlsx":
        yield from _iter_xlsx_rows(path)
        return
    if suffix == ".xls":
        yield from _iter_xls_rows(path)
        return
    if suffix == ".json":
        yield from _iter_json_rows(path)
        return
    if suffix == ".jsonl":
        yield from _iter_jsonl_rows(path)
        return
    if suffix == ".xml":
        yield from _iter_xml_rows(path)
        return
    raise ValueError(f"Unsupported file type: {suffix or '<none>'}. Supported: .csv, .tsv, .txt, .xlsx, .xls, .json, .jsonl, .xml")


def load_uploaded_headers(file_path: str | Path) -> list[str]:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_delimited_headers(path, ",")
    if suffix == ".tsv":
        return _read_delimited_headers(path, "\t")
    if suffix == ".txt":
        return _read_delimited_headers(path, ",")
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise ValueError("XLSX 업로드를 처리하려면 openpyxl이 필요합니다.") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        header_row = next(worksheet.iter_rows(values_only=True), None)
        workbook.close()
        return [header for header in _clean_headers(list(header_row or [])) if header]
    if suffix == ".xls":
        try:
            import xlrd
        except ImportError as exc:  # pragma: no cover
            raise ValueError("XLS 업로드를 처리하려면 xlrd가 필요합니다.") from exc
        workbook = xlrd.open_workbook(path)
        sheet = workbook.sheet_by_index(0)
        if sheet.nrows == 0:
            return []
        return [header for header in _clean_headers(sheet.row_values(0)) if header]
    if suffix == ".json":
        return _headers_from_rows(list(_iter_json_rows(path)))
    if suffix == ".jsonl":
        return _headers_from_rows(list(_iter_jsonl_rows(path)))
    if suffix == ".xml":
        return _headers_from_rows(list(_iter_xml_rows(path)))
    raise ValueError(f"Unsupported file type: {suffix or '<none>'}. Supported: .csv, .tsv, .txt, .xlsx, .xls, .json, .jsonl, .xml")


def _split_csv_list(value: str) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in value.split(",") if item.strip()]


def _file_cache_key(file_path: str | Path) -> tuple[str, int]:
    path = Path(file_path).resolve()
    return str(path), path.stat().st_mtime_ns


def load_dataset_meta(
    csv_path: str | Path,
    dataset_id: str | None = None,
    dataset_name: str | None = None,
) -> DatasetMeta:
    path = Path(csv_path)
    with _open_text_dataset(path) as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            row_id = (row.get("목록키") or "").strip()
            row_name = (row.get("목록명") or "").strip()
            if dataset_id and row_id != dataset_id:
                continue
            if dataset_name and row_name != dataset_name:
                continue
            total_rows = (row.get("전체행") or "").strip()
            return DatasetMeta(
                dataset_id=row_id,
                dataset_name=row_name,
                keywords=_split_csv_list(row.get("키워드", "")),
                provider_name=(row.get("제공기관명") or "").strip(),
                provider_code=(row.get("제공기관코드") or "").strip(),
                dataset_type=(row.get("목록유형") or "").strip(),
                service_type=(row.get("서비스 유형") or "").strip(),
                data_format=(row.get("데이터포맷") or "").strip(),
                request_fields=_split_csv_list(row.get("요청변수", "")),
                response_fields=_split_csv_list(row.get("출력결과", "")),
                update_cycle=(row.get("주기") or "").strip(),
                total_rows=int(total_rows) if total_rows.isdigit() else None,
            )
    target = dataset_id or dataset_name or "<unknown>"
    raise ValueError(f"Dataset not found: {target}")


def load_uploaded_dataset_meta(file_path: str | Path, dataset_name: str | None = None) -> DatasetMeta:
    path = Path(file_path)
    suffix = path.suffix.lstrip(".").lower() or "csv"
    header = load_uploaded_headers(path)
    if not header:
        raise ValueError("Uploaded dataset has no header row.")

    name = dataset_name or path.stem
    return DatasetMeta(
        dataset_id=f"{UPLOAD_DATASET_ID_PREFIX}{path.stem}",
        dataset_name=name,
        keywords=[],
        provider_name=UPLOAD_PROVIDER_NAME,
        provider_code=UPLOAD_PROVIDER_CODE,
        dataset_type=UPLOAD_DATASET_TYPE,
        service_type=UPLOAD_SERVICE_TYPE,
        data_format=suffix,
        request_fields=[],
        response_fields=[column.strip() for column in header if column.strip()],
        update_cycle=UPLOAD_UPDATE_CYCLE,
        total_rows=None,
    )
