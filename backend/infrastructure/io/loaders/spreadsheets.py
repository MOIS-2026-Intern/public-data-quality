from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path

from .common import clean_headers, stringify_cell


def iter_xlsx_rows(path: Path) -> Iterator[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ValueError("XLSX 업로드를 처리하려면 openpyxl이 필요합니다.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    row_iter = worksheet.iter_rows(values_only=True)
    headers = clean_headers(list(next(row_iter, None) or []))
    if not any(headers):
        workbook.close()
        raise ValueError("Uploaded dataset has no header row.")

    try:
        for row in row_iter:
            values = list(row or [])
            yield {header: stringify_cell(values[index]) if index < len(values) else "" for index, header in enumerate(headers) if header}
    finally:
        workbook.close()


def xlsx_headers(path: Path) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ValueError("XLSX 업로드를 처리하려면 openpyxl이 필요합니다.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    header_row = next(worksheet.iter_rows(values_only=True), None)
    workbook.close()
    return [header for header in clean_headers(list(header_row or [])) if header]


def iter_xls_rows(path: Path) -> Iterator[dict[str, str]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover
        raise ValueError("XLS 업로드를 처리하려면 xlrd가 필요합니다.") from exc

    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(0)
    if sheet.nrows == 0:
        raise ValueError("Uploaded dataset has no header row.")
    headers = clean_headers(sheet.row_values(0))
    if not any(headers):
        raise ValueError("Uploaded dataset has no header row.")

    for row_index in range(1, sheet.nrows):
        values = sheet.row_values(row_index)
        yield {header: stringify_cell(values[index]) if index < len(values) else "" for index, header in enumerate(headers) if header}


def xls_headers(path: Path) -> list[str]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover
        raise ValueError("XLS 업로드를 처리하려면 xlrd가 필요합니다.") from exc

    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(0)
    if sheet.nrows == 0:
        return []
    return [header for header in clean_headers(sheet.row_values(0)) if header]
