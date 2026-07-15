from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.config.reporting import MAX_COMMENT_LENGTH, REPORT_EXTENSION

UNSAFE_FILENAME_RE = re.compile(r'[\x00-\x1f\x7f/\\<>:"|?*]+')
HEADER_FILL = PatternFill("solid", fgColor="EAF0F6")
ISSUE_FILL = PatternFill("solid", fgColor="FBE3DF")
HEADER_FONT = Font(bold=True, color="10221B")


def report_filename(dataset_name: str) -> str:
    safe_name = UNSAFE_FILENAME_RE.sub("_", dataset_name).strip(" ._") or "dataset"
    stem = Path(safe_name).stem if Path(safe_name).suffix else safe_name
    return f"{(stem.strip(' ._') or 'dataset')}{REPORT_EXTENSION}"


def add_issue_comment(sheet, *, row_index: int, column_index: int, messages: list[str]) -> None:
    cell = sheet.cell(row=row_index, column=column_index)
    cell.fill = ISSUE_FILL
    cell.comment = Comment(comment_text(messages), "LDQ")


def result_dataset_name(result: dict[str, Any]) -> str:
    return str(result.get("summary", {}).get("dataset_name") or "dataset")


def finding_current_value(finding: dict[str, Any], row_index: int | None) -> str:
    if row_index is None:
        return ""
    row_values = finding.get("row_values") or {}
    if isinstance(row_values, dict):
        return str(row_values.get(str(row_index)) or row_values.get(row_index) or "")
    return ""


def finding_validation_area(finding: dict[str, Any]) -> str:
    return str(finding.get("category_label") or "")


def result_findings(result: dict[str, Any], *, finding_type: str) -> list[dict[str, Any]]:
    return [
        finding
        for finding in result.get("findings", [])
        if finding.get("finding_type") == finding_type
    ]


def result_headers(result: dict[str, Any], validation_rows: list[dict[str, str]]) -> list[str]:
    headers = [str(header) for header in result.get("preview_headers") or [] if str(header)]
    seen = set(headers)
    for row in validation_rows:
        for header in row:
            if header not in seen:
                headers.append(header)
                seen.add(header)
    return headers


def finding_row_indexes(finding: dict[str, Any]) -> list[int]:
    indexes: list[int] = []
    for row_index in finding.get("row_indexes") or []:
        try:
            parsed = int(row_index)
        except (TypeError, ValueError):
            continue
        if parsed > 0:
            indexes.append(parsed)
    return indexes


def issue_cells_by_position(result: dict[str, Any]) -> set[tuple[int, str]]:
    return set(issue_messages_by_cell(result))


def issue_messages_by_cell(result: dict[str, Any]) -> dict[tuple[int, str], list[str]]:
    issue_map: dict[tuple[int, str], list[str]] = defaultdict(list)
    for finding in result.get("findings", []):
        if finding.get("finding_type") != "issue":
            continue
        column_name = str(finding.get("column_name") or "")
        if not column_name:
            continue
        for row_index in finding_row_indexes(finding):
            issue_map[(row_index, column_name)].append(finding_message(finding))
    return issue_map


def finding_message(finding: dict[str, Any]) -> str:
    parts = [
        str(finding.get("category_label") or "").strip(),
        str(finding.get("criterion_name") or "").strip(),
        str(finding.get("message") or "").strip(),
    ]
    llm_final_verification = str(finding.get("llm_final_verification") or "").strip()
    if llm_final_verification:
        parts.append(f"LLM 최종 검증: {llm_final_verification}")
    rule_id = str(finding.get("rule_id") or "").strip()
    message = " / ".join(part for part in parts if part)
    return f"{message} ({rule_id})" if rule_id else message


def comment_text(messages: list[str]) -> str:
    return "\n\n".join(f"- {message}" for message in messages if message)[:MAX_COMMENT_LENGTH]


def style_header_row(sheet, row_index: int) -> None:
    for cell in sheet[row_index]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")


def auto_width(sheet, max_width: int = 36) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        width = max(min(len(str(cell.value or "")) + 2, max_width) for cell in column_cells)
        sheet.column_dimensions[column_letter].width = max(width, 10)
