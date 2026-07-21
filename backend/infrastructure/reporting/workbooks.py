from __future__ import annotations

import re
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from backend.config.reporting import (
    BATCH_COLUMN_ERROR_REPORT_ARCHIVE_FILENAME,
    BATCH_COLUMN_ERROR_REPORT_FILENAME,
    BATCH_REPORT_FILENAME,
    REPORTS_DIR_NAME,
)
from .workbook_support import (
    add_issue_comment,
    auto_width as _auto_width,
    finding_current_value as _finding_current_value,
    finding_row_indexes as _finding_row_indexes,
    finding_validation_area as _finding_validation_area,
    issue_details_by_cell as _issue_details_by_cell,
    issue_cells_by_position as _issue_cells_by_position,
    issue_messages_by_cell as _issue_messages_by_cell,
    report_filename as _report_filename,
    result_dataset_name as _result_dataset_name,
    result_findings as _result_findings,
    result_headers as _result_headers,
    style_header_row as _style_header_row,
)
from .artifacts import unique_artifact_path

_SHEET_TITLE_UNSAFE_RE = re.compile(r"[\[\]:*?/\\]+")
_BATCH_COLUMN_ERROR_REPORT_CHUNK_SIZE = 30


def write_error_report(
    *,
    result: dict[str, Any],
    validation_rows: list[dict[str, str]],
    output_dir: Path,
) -> Path:
    reports_dir = output_dir / REPORTS_DIR_NAME
    reports_dir.mkdir(parents=True, exist_ok=True)

    dataset_name = str(result.get("summary", {}).get("dataset_name") or "dataset")
    report_path = unique_artifact_path(reports_dir, _report_filename(dataset_name))

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "요약"
    _write_summary_sheet(summary_sheet, result, validation_rows)
    _write_data_sheet(workbook.create_sheet("전체 데이터 오류 표시"), result, validation_rows)
    _write_column_error_sheet(workbook.create_sheet("컬럼별 데이터 오류 표시"), result, validation_rows)
    _write_findings_sheet(workbook.create_sheet("오류 상세"), result, validation_rows)
    _write_manual_review_sheet(workbook.create_sheet("수동 검토 상세"), result, validation_rows)

    workbook.save(report_path)
    return report_path


def write_batch_error_report(
    *,
    items: list[dict[str, Any]],
    output_dir: Path,
) -> Path:
    reports_dir = output_dir / REPORTS_DIR_NAME
    reports_dir.mkdir(parents=True, exist_ok=True)

    report_path = unique_artifact_path(reports_dir, BATCH_REPORT_FILENAME)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "요약"
    successful_results = _successful_batch_results(items)
    _write_batch_summary_sheet(summary_sheet, successful_results)
    _write_batch_findings_sheet(workbook.create_sheet("오류 상세"), successful_results)
    _write_batch_manual_review_sheet(workbook.create_sheet("수동 검토 상세"), successful_results)

    workbook.save(report_path)
    return report_path


def write_batch_column_error_report_archive(
    *,
    report_paths: list[Path],
    output_dir: Path,
) -> Path:
    reports_dir = output_dir / REPORTS_DIR_NAME
    reports_dir.mkdir(parents=True, exist_ok=True)

    archive_path = unique_artifact_path(reports_dir, BATCH_COLUMN_ERROR_REPORT_ARCHIVE_FILENAME)
    used_names: set[str] = set()
    with zipfile.ZipFile(archive_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for report_path in report_paths:
            archive.write(report_path, arcname=_unique_archive_member_name(report_path.name, used_names))
    return archive_path


def write_batch_column_error_report(
    *,
    entries: list[dict[str, Any]],
    output_dir: Path,
) -> list[Path]:
    reports_dir = output_dir / REPORTS_DIR_NAME
    reports_dir.mkdir(parents=True, exist_ok=True)

    successful_entries = [entry for entry in entries if isinstance(entry.get("result"), dict)]
    if not successful_entries:
        report_path = unique_artifact_path(reports_dir, BATCH_COLUMN_ERROR_REPORT_FILENAME)
        workbook = Workbook()
        workbook.active.title = "데이터"
        workbook.save(report_path)
        return [report_path]

    report_groups = _batch_column_error_report_groups(successful_entries)
    ordinary_group_count = sum(1 for group in report_groups if group["kind"] == "ordinary")
    report_paths: list[Path] = []
    for group in report_groups:
        report_path = unique_artifact_path(
            reports_dir,
            _batch_column_error_report_group_filename(group, ordinary_group_count),
        )
        workbook = _build_column_error_workbook(
            group["entries"],
            prefer_archive_member_sheet_title=group["kind"] == "archive",
        )
        workbook.save(report_path)
        report_paths.append(report_path)
    return report_paths


def _batch_column_error_report_groups(entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    archive_infos = [_archive_source_info(entry) for entry in entries]
    archive_counts: dict[str, int] = {}
    for archive_info in archive_infos:
        if archive_info is None:
            continue
        archive_name, _ = archive_info
        archive_counts[archive_name] = archive_counts.get(archive_name, 0) + 1

    ordinary_entries: list[dict[str, Any]] = []
    archive_entries_by_name: dict[str, list[dict[str, Any]]] = {}
    for entry, archive_info in zip(entries, archive_infos, strict=False):
        if archive_info is None or archive_counts[archive_info[0]] <= 1:
            ordinary_entries.append(entry)
            continue
        archive_entries_by_name.setdefault(archive_info[0], []).append(entry)

    groups: list[dict[str, Any]] = [
        {"kind": "ordinary", "chunk_index": index, "entries": entry_chunk}
        for index, entry_chunk in enumerate(
            _chunks(ordinary_entries, _BATCH_COLUMN_ERROR_REPORT_CHUNK_SIZE),
            start=1,
        )
    ]
    groups.extend(
        {"kind": "archive", "archive_name": archive_name, "entries": archive_entries}
        for archive_name, archive_entries in archive_entries_by_name.items()
    )
    return groups


def _build_column_error_workbook(
    entries: list[dict[str, Any]],
    *,
    prefer_archive_member_sheet_title: bool = False,
) -> Workbook:
    workbook = Workbook()
    used_titles: set[str] = set()
    first_entry = entries[0]
    first_sheet = workbook.active
    first_sheet.title = _unique_sheet_title(
        _entry_sheet_title(first_entry, prefer_archive_member=prefer_archive_member_sheet_title),
        used_titles,
    )
    _write_column_error_sheet(
        first_sheet,
        first_entry["result"],
        list(first_entry.get("validation_rows") or []),
    )

    for entry in entries[1:]:
        sheet = workbook.create_sheet(
            _unique_sheet_title(
                _entry_sheet_title(entry, prefer_archive_member=prefer_archive_member_sheet_title),
                used_titles,
            )
        )
        _write_column_error_sheet(
            sheet,
            entry["result"],
            list(entry.get("validation_rows") or []),
        )
    return workbook


def _batch_column_error_report_group_filename(group: dict[str, Any], ordinary_group_count: int) -> str:
    if group["kind"] == "archive":
        return _archive_column_error_report_filename(str(group.get("archive_name") or "archive.zip"))
    if ordinary_group_count > 1:
        return _batch_column_error_report_filename(int(group.get("chunk_index") or 1))
    return BATCH_COLUMN_ERROR_REPORT_FILENAME


def _batch_column_error_report_filename(chunk_index: int) -> str:
    path = Path(BATCH_COLUMN_ERROR_REPORT_FILENAME)
    return f"{path.stem}_{chunk_index:02d}{path.suffix}"


def _unique_archive_member_name(filename: str, used_names: set[str]) -> str:
    source = Path(filename)
    candidate = source.name or "report.xlsx"
    if candidate not in used_names:
        used_names.add(candidate)
        return candidate

    stem = source.stem or "report"
    suffix = source.suffix
    index = 2
    while True:
        candidate = f"{stem}_{index}{suffix}"
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        index += 1


def _archive_column_error_report_filename(archive_name: str) -> str:
    archive_stem = Path(str(archive_name).replace("\\", "/")).stem or "archive"
    return _report_filename(f"{archive_stem}_컬럼별_데이터_오류.xlsx")


def _chunks(items: list[dict[str, Any]], chunk_size: int) -> list[list[dict[str, Any]]]:
    return [items[index : index + chunk_size] for index in range(0, len(items), chunk_size)]


def _entry_display_name(entry: dict[str, Any]) -> str:
    return str(
        entry.get("source_display_name")
        or entry.get("display_name")
        or _result_dataset_name(entry["result"])
    )


def _entry_sheet_title(entry: dict[str, Any], *, prefer_archive_member: bool) -> str:
    if prefer_archive_member:
        archive_info = _archive_source_info(entry)
        if archive_info is not None:
            _, member_name = archive_info
            return member_name
    return _result_dataset_name(entry["result"])


def _archive_source_info(entry: dict[str, Any]) -> tuple[str, str] | None:
    parts = [part for part in _entry_display_name(entry).replace("\\", "/").split("/") if part]
    for index, part in enumerate(parts[:-1]):
        if Path(part).suffix.lower() == ".zip":
            return "/".join(parts[: index + 1]), "/".join(parts[index + 1 :]) or parts[-1]
    return None


def _write_summary_sheet(sheet, result: dict[str, Any], validation_rows: list[dict[str, str]]) -> None:
    summary = result.get("summary", {})
    issue_cells = _issue_cells_by_position(result)
    issue_rows = {row_index for row_index, _ in issue_cells}
    issue_columns = {column_name for _, column_name in issue_cells}
    rows = [
        ("전체 컬럼 수", summary.get("column_count", 0)),
        ("전체 행 수", summary.get("row_count") or len(validation_rows)),
        ("전체 오류 발생 컬럼 수", len(issue_columns)),
        ("오류 발생 행 수", len(issue_rows)),
    ]
    sheet.append(["항목", "값"])
    for row in rows:
        sheet.append(list(row))
    _style_header_row(sheet, 1)
    _auto_width(sheet)
    sheet.freeze_panes = "A2"


def _write_data_sheet(sheet, result: dict[str, Any], validation_rows: list[dict[str, str]]) -> None:
    headers = _result_headers(result, validation_rows)
    sheet.append(["row_index", *headers])
    _style_header_row(sheet, 1)
    sheet.freeze_panes = "A2"

    issue_map = _issue_messages_by_cell(result)
    max_finding_row = max((row_index for row_index, _ in issue_map), default=0)
    row_count = max(len(validation_rows), max_finding_row)

    for row_index in range(1, row_count + 1):
        source_row = validation_rows[row_index - 1] if row_index <= len(validation_rows) else {}
        sheet.append([row_index, *[source_row.get(header, "") for header in headers]])

    column_index_by_name = {header: index + 2 for index, header in enumerate(headers)}
    for (row_index, column_name), messages in issue_map.items():
        excel_row = row_index + 1
        excel_column = column_index_by_name.get(column_name)
        if not excel_column:
            continue
        add_issue_comment(sheet, row_index=excel_row, column_index=excel_column, messages=messages)

    _auto_width(sheet, max_width=42)


def _write_column_error_sheet(sheet, result: dict[str, Any], validation_rows: list[dict[str, str]]) -> None:
    headers = _result_headers(result, validation_rows)
    display_headers = ["row_index", *headers, "오류 여부", "오류 내용"]

    sheet.append(display_headers)
    _style_header_row(sheet, 1)
    sheet.freeze_panes = "A2"

    issue_map = _issue_details_by_cell(result)
    row_issue_entries: dict[int, list[tuple[str, str]]] = {}
    for (row_index, column_name), details in issue_map.items():
        filtered_details = [detail for detail in details if detail]
        if not filtered_details:
            continue
        row_issue_entries.setdefault(row_index, []).extend((column_name, detail) for detail in filtered_details)

    max_finding_row = max(row_issue_entries, default=0)
    row_count = max(len(validation_rows), max_finding_row)

    for row_index in range(1, row_count + 1):
        source_row = validation_rows[row_index - 1] if row_index <= len(validation_rows) else {}
        row_entries = row_issue_entries.get(row_index, [])
        row_details = [
            detail if len(row_entries) == 1 or not column_name else f"{column_name}: {detail}"
            for column_name, detail in row_entries
        ]
        sheet.append(
            [
                row_index,
                *[source_row.get(header, "") for header in headers],
                "오류" if row_details else "",
                " / ".join(row_details),
            ]
        )

    _apply_error_status_filter(sheet, display_headers)
    _auto_width(sheet, max_width=52)


def _apply_error_status_filter(sheet, headers: list[str]) -> None:
    try:
        error_status_column_index = headers.index("오류 여부") + 1
    except ValueError:
        return
    if sheet.max_row < 2:
        return

    column_letter = get_column_letter(error_status_column_index)
    sheet.auto_filter.ref = f"{column_letter}1:{column_letter}{sheet.max_row}"
    sheet.auto_filter.add_filter_column(0, ["오류"])


def _write_findings_sheet(sheet, result: dict[str, Any], validation_rows: list[dict[str, str]]) -> None:
    headers = [
        "데이터명",
        "컬럼명",
        "행 번호",
        "검증영역",
        "현재 값",
        "오류 메세지",
        "LLM 최종 검증",
    ]
    sheet.append(headers)
    _style_header_row(sheet, 1)
    sheet.freeze_panes = "A2"

    dataset_name = str(result.get("summary", {}).get("dataset_name") or "")
    for finding in _result_findings(result, finding_type="issue"):
        column_name = str(finding.get("column_name") or "")
        row_indexes = _finding_row_indexes(finding)
        if not row_indexes:
            row_indexes = [None]
        for row_index in row_indexes:
            current_value = _row_current_value(
                finding,
                row_index=row_index,
                validation_rows=validation_rows,
                column_name=column_name,
            )
            sheet.append(
                [
                    dataset_name,
                    column_name,
                    row_index or "",
                    _finding_validation_area(finding),
                    current_value,
                    finding.get("message", ""),
                    finding.get("llm_final_verification", ""),
                ]
            )

    _auto_width(sheet, max_width=48)


def _write_manual_review_sheet(sheet, result: dict[str, Any], validation_rows: list[dict[str, str]]) -> None:
    headers = [
        "데이터명",
        "컬럼명",
        "행 번호",
        "검증영역",
        "현재 값",
        "오류 메세지",
    ]
    sheet.append(headers)
    _style_header_row(sheet, 1)
    sheet.freeze_panes = "A2"

    dataset_name = str(result.get("summary", {}).get("dataset_name") or "")
    for finding in _result_findings(result, finding_type="manual_review"):
        column_name = str(finding.get("column_name") or "")
        row_indexes = _finding_row_indexes(finding) or [None]
        for row_index in row_indexes:
            sheet.append(
                [
                    dataset_name,
                    column_name,
                    row_index or "",
                    _finding_validation_area(finding),
                    _row_current_value(
                        finding,
                        row_index=row_index,
                        validation_rows=validation_rows,
                        column_name=column_name,
                    ),
                    finding.get("message", ""),
                ]
            )

    _auto_width(sheet, max_width=48)


def _write_batch_summary_sheet(sheet, results: list[dict[str, Any]]) -> None:
    issue_cells = {
        (dataset_name, row_index, column_name)
        for result in results
        for dataset_name in [_result_dataset_name(result)]
        for row_index, column_name in _issue_cells_by_position(result)
    }
    issue_rows = {(dataset_name, row_index) for dataset_name, row_index, _ in issue_cells}
    issue_columns = {(dataset_name, column_name) for dataset_name, _, column_name in issue_cells}
    rows = [
        ("전체 데이터 개수", len(results)),
        ("전체 컬럼 수", sum(int(result.get("summary", {}).get("column_count") or 0) for result in results)),
        ("전체 행 수", sum(int(result.get("summary", {}).get("row_count") or 0) for result in results)),
        ("전체 오류 발생 컬럼 수", len(issue_columns)),
        ("전체 오류 발생 행 수", len(issue_rows)),
    ]
    sheet.append(["항목", "값"])
    for row in rows:
        sheet.append(list(row))
    _style_header_row(sheet, 1)
    _auto_width(sheet)
    sheet.freeze_panes = "A2"


def _write_batch_findings_sheet(sheet, results: list[dict[str, Any]]) -> None:
    headers = [
        "데이터명",
        "컬럼명",
        "행 번호",
        "검증영역",
        "현재 값",
        "오류 메세지",
        "LLM 최종 검증",
    ]
    sheet.append(headers)
    _style_header_row(sheet, 1)
    sheet.freeze_panes = "A2"

    for result in results:
        dataset_name = _result_dataset_name(result)
        for finding in _result_findings(result, finding_type="issue"):
            column_name = str(finding.get("column_name") or "")
            row_indexes = _finding_row_indexes(finding) or [None]
            for row_index in row_indexes:
                sheet.append(
                    [
                        dataset_name,
                        column_name,
                        row_index or "",
                        _finding_validation_area(finding),
                        _finding_current_value(finding, row_index),
                        finding.get("message", ""),
                        finding.get("llm_final_verification", ""),
                    ]
                )
    _auto_width(sheet, max_width=48)


def _write_batch_manual_review_sheet(sheet, results: list[dict[str, Any]]) -> None:
    headers = [
        "데이터명",
        "컬럼명",
        "행 번호",
        "검증영역",
        "현재 값",
        "오류 메세지",
    ]
    sheet.append(headers)
    _style_header_row(sheet, 1)
    sheet.freeze_panes = "A2"

    for result in results:
        dataset_name = _result_dataset_name(result)
        for finding in _result_findings(result, finding_type="manual_review"):
            column_name = str(finding.get("column_name") or "")
            row_indexes = _finding_row_indexes(finding) or [None]
            for row_index in row_indexes:
                sheet.append(
                    [
                        dataset_name,
                        column_name,
                        row_index or "",
                        _finding_validation_area(finding),
                        _finding_current_value(finding, row_index),
                        finding.get("message", ""),
                    ]
                )
    _auto_width(sheet, max_width=48)


def _row_current_value(
    finding: dict[str, Any],
    *,
    row_index: int | None,
    validation_rows: list[dict[str, str]],
    column_name: str,
) -> str:
    current_value = _finding_current_value(finding, row_index)
    if current_value or row_index is None:
        return current_value
    if 0 < row_index <= len(validation_rows):
        return str(validation_rows[row_index - 1].get(column_name, ""))
    return ""


def _unique_sheet_title(title: str, used_titles: set[str]) -> str:
    sanitized = _SHEET_TITLE_UNSAFE_RE.sub("_", str(title or "")).strip(" '") or "dataset"
    candidate = sanitized[:31] or "dataset"
    if candidate not in used_titles:
        used_titles.add(candidate)
        return candidate

    suffix = 2
    while True:
        suffix_text = f"_{suffix}"
        trimmed = (sanitized[: 31 - len(suffix_text)] or "dataset") + suffix_text
        if trimmed not in used_titles:
            used_titles.add(trimmed)
            return trimmed
        suffix += 1


def _successful_batch_results(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    for item in items:
        if not item.get("ok") or not isinstance(item.get("result"), dict):
            continue
        result = item["result"]
        if not result.get("summary", {}).get("dataset_name"):
            result = {
                **result,
                "summary": {
                    **result.get("summary", {}),
                    "dataset_name": item.get("filename") or "dataset",
                },
            }
        results.append(result)
    return results
