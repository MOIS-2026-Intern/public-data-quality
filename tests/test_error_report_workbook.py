from pathlib import Path
import sys

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from backend.infrastructure.reporting.workbooks import (
    write_batch_column_error_report,
    write_batch_error_report,
    write_error_report,
)


def _issue_finding(column_name: str, row_index: int, value: str, message: str) -> dict:
    return {
        "column_name": column_name,
        "finding_type": "issue",
        "category_label": "컬럼 완결성 검증",
        "row_indexes": [row_index],
        "message": message,
        "llm_final_verification": "LLM 확인 결과 오류입니다.",
        "row_values": {str(row_index): value},
    }


def _manual_review_finding(column_name: str, row_index: int, value: str, message: str) -> dict:
    return {
        "column_name": column_name,
        "finding_type": "manual_review",
        "category_label": "컬럼 특성 유효성 검증",
        "row_indexes": [row_index],
        "message": message,
        "row_values": {str(row_index): value},
    }


def _assert_error_status_filter(sheet, ref: str) -> None:
    assert sheet.auto_filter.ref == ref
    assert len(sheet.auto_filter.filterColumn) == 1
    filter_column = sheet.auto_filter.filterColumn[0]
    assert filter_column.colId == 0
    assert list(filter_column.filters.filter) == ["오류"]
    assert filter_column.filters.blank is False


def test_single_error_report_uses_requested_sheets_and_detail_columns(tmp_path) -> None:
    result = {
        "summary": {"dataset_name": "화성시_어린이보호구역.csv", "column_count": 2, "row_count": 2},
        "preview_headers": ["시설명", "가격"],
        "findings": [
            _issue_finding("가격", 2, "메뉴삭제", "금액 컬럼에 금액이 아닌 값이 있습니다."),
            _manual_review_finding("시설명", 1, "A", "시설명 값은 의미 판정이 애매해 수동 검토가 필요합니다."),
        ],
    }
    validation_rows = [
        {"시설명": "A", "가격": "1000"},
        {"시설명": "B", "가격": "메뉴삭제"},
    ]

    report_path = write_error_report(result=result, validation_rows=validation_rows, output_dir=tmp_path)
    workbook = load_workbook(report_path)

    assert report_path.suffix == ".xlsx"
    assert report_path.name == "화성시_어린이보호구역.xlsx"
    assert workbook.sheetnames == ["요약", "전체 데이터 오류 표시", "컬럼별 데이터 오류 표시", "오류 상세", "수동 검토 상세"]
    assert [cell.value for cell in workbook["요약"]["A"]] == [
        "항목",
        "전체 컬럼 수",
        "전체 행 수",
        "전체 오류 발생 컬럼 수",
        "오류 발생 행 수",
    ]
    assert [cell.value for cell in workbook["컬럼별 데이터 오류 표시"][1]] == [
        "row_index",
        "시설명",
        "가격",
        "오류 여부",
        "오류 내용",
    ]
    assert [cell.value for cell in workbook["컬럼별 데이터 오류 표시"][3]] == [
        2,
        "B",
        "메뉴삭제",
        "오류",
        "금액 컬럼에 금액이 아닌 값이 있습니다., LLM 확인 결과 오류입니다.",
    ]
    _assert_error_status_filter(workbook["컬럼별 데이터 오류 표시"], "D1:D3")
    assert [cell.value for cell in workbook["오류 상세"][1]] == [
        "데이터명",
        "컬럼명",
        "행 번호",
        "검증영역",
        "현재 값",
        "오류 메세지",
        "LLM 최종 검증",
    ]
    assert [cell.value for cell in workbook["오류 상세"][2]] == [
        "화성시_어린이보호구역.csv",
        "가격",
        2,
        "컬럼 완결성 검증",
        "메뉴삭제",
        "금액 컬럼에 금액이 아닌 값이 있습니다.",
        "LLM 확인 결과 오류입니다.",
    ]
    assert [cell.value for cell in workbook["수동 검토 상세"][1]] == [
        "데이터명",
        "컬럼명",
        "행 번호",
        "검증영역",
        "현재 값",
        "오류 메세지",
    ]
    assert [cell.value for cell in workbook["수동 검토 상세"][2]] == [
        "화성시_어린이보호구역.csv",
        "시설명",
        1,
        "컬럼 특성 유효성 검증",
        "A",
        "시설명 값은 의미 판정이 애매해 수동 검토가 필요합니다.",
    ]


def test_batch_error_report_has_summary_and_detail_sheets(tmp_path) -> None:
    items = [
        {
            "ok": True,
            "filename": "A.csv",
            "result": {
                "summary": {"dataset_name": "A.csv", "column_count": 2, "row_count": 2},
                "findings": [
                    _issue_finding("가격", 2, "메뉴삭제", "금액 오류"),
                    _manual_review_finding("시설명", 1, "A", "시설명은 수동 검토가 필요합니다."),
                ],
            },
        },
        {
            "ok": True,
            "filename": "B.csv",
            "result": {
                "summary": {"dataset_name": "B.csv", "column_count": 1, "row_count": 1},
                "findings": [_issue_finding("연락처", 1, "abc", "연락처 오류")],
            },
        },
    ]

    report_path = write_batch_error_report(items=items, output_dir=tmp_path)
    workbook = load_workbook(report_path)

    assert report_path.suffix == ".xlsx"
    assert report_path.name == "전체_오류_리포트.xlsx"
    assert workbook.sheetnames == ["요약", "오류 상세", "수동 검토 상세"]
    summary_values = {row[0].value: row[1].value for row in workbook["요약"].iter_rows(min_row=2, max_col=2)}
    assert summary_values == {
        "전체 데이터 개수": 2,
        "전체 컬럼 수": 3,
        "전체 행 수": 3,
        "전체 오류 발생 컬럼 수": 2,
        "전체 오류 발생 행 수": 2,
    }
    assert [cell.value for cell in workbook["오류 상세"][2]] == [
        "A.csv",
        "가격",
        2,
        "컬럼 완결성 검증",
        "메뉴삭제",
        "금액 오류",
        "LLM 확인 결과 오류입니다.",
    ]
    assert [cell.value for cell in workbook["수동 검토 상세"][1]] == [
        "데이터명",
        "컬럼명",
        "행 번호",
        "검증영역",
        "현재 값",
        "오류 메세지",
    ]
    assert [cell.value for cell in workbook["수동 검토 상세"][2]] == [
        "A.csv",
        "시설명",
        1,
        "컬럼 특성 유효성 검증",
        "A",
        "시설명은 수동 검토가 필요합니다.",
    ]


def test_batch_column_error_report_creates_one_sheet_per_dataset(tmp_path) -> None:
    entries = [
        {
            "result": {
                "summary": {"dataset_name": "A.csv", "column_count": 2, "row_count": 2},
                "preview_headers": ["시설명", "가격"],
                "findings": [_issue_finding("가격", 2, "메뉴삭제", "금액 오류")],
            },
            "validation_rows": [
                {"시설명": "A", "가격": "1000"},
                {"시설명": "B", "가격": "메뉴삭제"},
            ],
        },
        {
            "result": {
                "summary": {"dataset_name": "B.csv", "column_count": 1, "row_count": 1},
                "preview_headers": ["연락처"],
                "findings": [_issue_finding("연락처", 1, "abc", "연락처 오류")],
            },
            "validation_rows": [{"연락처": "abc"}],
        },
    ]

    report_paths = write_batch_column_error_report(entries=entries, output_dir=tmp_path)
    assert [path.name for path in report_paths] == ["A.xlsx", "B.xlsx"]
    first_workbook = load_workbook(report_paths[0])
    second_workbook = load_workbook(report_paths[1])

    assert first_workbook.sheetnames == ["A.csv"]
    assert second_workbook.sheetnames == ["B.csv"]
    assert [cell.value for cell in first_workbook["A.csv"][1]] == [
        "row_index",
        "시설명",
        "가격",
        "오류 여부",
        "오류 내용",
    ]
    assert [cell.value for cell in first_workbook["A.csv"][3]] == [
        2,
        "B",
        "메뉴삭제",
        "오류",
        "금액 오류, LLM 확인 결과 오류입니다.",
    ]
    assert [cell.value for cell in second_workbook["B.csv"][2]] == [
        1,
        "abc",
        "오류",
        "연락처 오류, LLM 확인 결과 오류입니다.",
    ]
    _assert_error_status_filter(first_workbook["A.csv"], "D1:D3")
    _assert_error_status_filter(second_workbook["B.csv"], "C1:C2")


def test_batch_column_error_report_creates_individual_files_for_regular_datasets(tmp_path) -> None:
    entries = [
        {
            "result": {
                "summary": {"dataset_name": f"data-{index:02d}.csv", "column_count": 1, "row_count": 1},
                "preview_headers": ["value"],
                "findings": [_issue_finding("value", 1, str(index), "값 오류")],
            },
            "validation_rows": [{"value": str(index)}],
        }
        for index in range(1, 32)
    ]

    report_paths = write_batch_column_error_report(entries=entries, output_dir=tmp_path)

    assert [path.name for path in report_paths] == [f"data-{index:02d}.xlsx" for index in range(1, 32)]
    first_workbook = load_workbook(report_paths[0])
    last_workbook = load_workbook(report_paths[-1])
    assert first_workbook.sheetnames == ["data-01.csv"]
    assert last_workbook.sheetnames == ["data-31.csv"]


def test_batch_column_error_report_keeps_zip_members_in_dedicated_workbook(tmp_path) -> None:
    ordinary_entries = [
        {
            "result": {
                "summary": {"dataset_name": f"normal-{index:02d}.csv", "column_count": 1, "row_count": 1},
                "preview_headers": ["value"],
                "findings": [_issue_finding("value", 1, str(index), "값 오류")],
            },
            "validation_rows": [{"value": str(index)}],
        }
        for index in range(1, 31)
    ]
    zip_entries = [
        {
            "result": {
                "summary": {"dataset_name": f"archive.zip/inner-{index}.csv", "column_count": 1, "row_count": 1},
                "preview_headers": ["value"],
                "findings": [_issue_finding("value", 1, str(index), "값 오류")],
            },
            "validation_rows": [{"value": str(index)}],
            "source_display_name": f"archive.zip/inner-{index}.csv",
        }
        for index in range(1, 3)
    ]

    report_paths = write_batch_column_error_report(
        entries=[*ordinary_entries[:29], *zip_entries, ordinary_entries[29]],
        output_dir=tmp_path,
    )

    assert [path.name for path in report_paths] == [
        *[f"normal-{index:02d}.xlsx" for index in range(1, 31)],
        "archive.xlsx",
    ]
    ordinary_workbook = load_workbook(report_paths[0])
    zip_workbook = load_workbook(report_paths[-1])
    assert ordinary_workbook.sheetnames == ["normal-01.csv"]
    assert zip_workbook.sheetnames == ["inner-1.csv", "inner-2.csv"]


def test_column_error_report_aggregates_multiple_column_errors_on_same_row(tmp_path) -> None:
    result = {
        "summary": {"dataset_name": "sample.csv", "column_count": 2, "row_count": 1},
        "preview_headers": ["주택보험금(천원)", "온실지급건수(건)"],
        "findings": [
            _issue_finding("주택보험금(천원)", 1, "", "주택보험금 오류"),
            _issue_finding("온실지급건수(건)", 1, "1", "온실지급건수 오류"),
        ],
    }
    validation_rows = [{"주택보험금(천원)": "", "온실지급건수(건)": "1"}]

    report_path = write_error_report(result=result, validation_rows=validation_rows, output_dir=tmp_path)
    workbook = load_workbook(report_path)

    assert [cell.value for cell in workbook["컬럼별 데이터 오류 표시"][2]] == [
        1,
        None,
        "1",
        "오류",
        (
            "주택보험금(천원): 주택보험금 오류, LLM 확인 결과 오류입니다. / "
            "온실지급건수(건): 온실지급건수 오류, LLM 확인 결과 오류입니다."
        ),
    ]
